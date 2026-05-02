import warnings
import zipfile
from xml.etree import ElementTree as ET

import openpyxl
from oletools.olevba import VBA_Parser, filter_vba
from openpyxl.worksheet.worksheet import Worksheet


def _get_used_bounds(ws):
    min_r, max_r, min_c, max_c = None, None, None, None
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if val is not None and str(val).strip():
                r, c = cell.row, cell.column
                if min_r is None or r < min_r:
                    min_r = r
                if max_r is None or r > max_r:
                    max_r = r
                if min_c is None or c < min_c:
                    min_c = c
                if max_c is None or c > max_c:
                    max_c = c
    if min_r is None:
        return None
    return min_r, max_r, min_c, max_c


def _extract_vba(filepath: str) -> str:
    parts = []
    try:
        with zipfile.ZipFile(filepath, "r") as zf:
            if "xl/vbaProject.bin" not in zf.namelist():
                return ""
            zf.read("xl/vbaProject.bin")
    except (zipfile.BadZipFile, KeyError):
        return ""

    try:
        vba_parser = VBA_Parser(filepath)
        if not vba_parser.detect_vba_macros():
            return ""
        macros = vba_parser.extract_all_macros()
        if not macros:
            return ""

        parts.append("\n## VBA Macros\n")
        seen_modules = set()
        for _, stream_path, vba_filename, vba_code in macros:
            key = (vba_filename, stream_path)
            if key in seen_modules:
                continue
            seen_modules.add(key)
            filtered = filter_vba(vba_code)
            if filtered.strip():
                parts.append(f"### `{vba_filename}` ({stream_path})\n")
                parts.append(f"```vba\n{filtered.strip()}\n```\n")
    except Exception:
        pass

    return "\n".join(parts)


def _extract_shapes_from_vml(vml_xml: str) -> list:
    shapes = []
    try:
        root = ET.fromstring(vml_xml)
    except ET.ParseError:
        return shapes

    for shape in root.iter("{urn:schemas-microsoft-com:vml}shape"):
        info = {}
        style = shape.get("style", "")
        info["name"] = shape.get("id", "")
        info["type"] = shape.get("type", "")
        info["style"] = style

        textbox = shape.find("{urn:schemas-microsoft-com:vml}textbox")
        if textbox is not None:
            text_content = "".join(textbox.itertext()).strip()
            if text_content:
                info["text"] = text_content

        anchor = shape.find("{urn:schemas-microsoft-com:office:excel}ClientData")
        if anchor is not None:
            obj_type = anchor.find("{urn:schemas-microsoft-com:office:excel}ObjectType")
            if obj_type is not None:
                info["object_type"] = obj_type.text
            macro = anchor.find("{urn:schemas-microsoft-com:office:excel}FmlaMacro")
            if macro is not None:
                info["macro"] = macro.text
            anchor_text = anchor.find("{urn:schemas-microsoft-com:office:excel}Anchor")
            if anchor_text is not None and anchor_text.text:
                info["cell_anchor"] = anchor_text.text.strip()

        shapes.append(info)
    return shapes


def _extract_shapes_from_drawing(drawing_xml: str) -> list:
    shapes = []
    try:
        root = ET.fromstring(drawing_xml)
    except ET.ParseError:
        return shapes

    xdr = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"
    a = "{http://schemas.openxmlformats.org/drawingml/2006/main}"

    for sp in root.iter(f"{xdr}sp"):
        info = {}
        macro_attr = sp.get("macro", "")
        if macro_attr:
            info["macro"] = macro_attr
        nv = sp.find(f"{xdr}nvSpPr")
        if nv is not None:
            c_nv = nv.find(f"{xdr}cNvPr")
            if c_nv is not None:
                info["name"] = c_nv.get("name", "")

        tx_body = sp.find(f"{xdr}txBody")
        if tx_body is not None:
            texts = []
            for t_elem in tx_body.iter(f"{a}t"):
                if t_elem.text:
                    texts.append(t_elem.text)
            if texts:
                info["text"] = " ".join(texts)

        anchor = sp.find(f"{xdr}clientData")
        if anchor is not None:
            info["fPrintsWithSheet"] = anchor.get("fPrintsWithSheet", "1")

        shapes.append(info)
    return shapes


def _extract_images(filepath: str) -> list:
    images = []
    try:
        with zipfile.ZipFile(filepath, "r") as zf:
            for name in zf.namelist():
                if name.startswith("xl/media/"):
                    data = zf.read(name)
                    images.append({
                        "path": name,
                        "size_bytes": len(data),
                        "extension": name.rsplit(".", 1)[-1] if "." in name else "unknown",
                    })
    except (zipfile.BadZipFile, KeyError):
        pass
    return images


def convert_xlsx(filepath: str) -> str:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        wb_values = openpyxl.load_workbook(filepath, data_only=True)
        wb_formulas = openpyxl.load_workbook(filepath, data_only=False)

    parts = []

    for sheet_name in wb_formulas.sheetnames:
        ws = wb_formulas[sheet_name]
        ws_vals = wb_values[sheet_name]
        sheet_parts = []
        sheet_parts.append(f"## {sheet_name}\n")

        if not isinstance(ws, Worksheet):
            sheet_parts.append("_(chartsheet or diagramsheet — no tabular data)_")
            parts.append("\n".join(sheet_parts))
            continue

        bounds = _get_used_bounds(ws)
        if bounds is None:
            sheet_parts.append("_(empty sheet)_")
            parts.append("\n".join(sheet_parts))
            continue

        min_r, max_r, min_c, max_c = bounds
        rows = []
        for r in range(min_r, max_r + 1):
            cells = []
            for c in range(min_c, max_c + 1):
                cell = ws.cell(row=r, column=c)
                val = ws_vals.cell(row=r, column=c).value
                formula = cell.value
                if formula and isinstance(formula, str) and formula.startswith("="):
                    val_str = str(val) if val is not None else "#N/A"
                    display_formula = formula.replace("_xlfn.", "")
                    cells.append(f"{val_str}\n`{display_formula}`")
                else:
                    cells.append(str(val) if val is not None else "")
            line = "| " + " | ".join(c.replace("|", "\\|").replace("\n", " ") for c in cells) + " |"
            rows.append(line)
            if r == min_r:
                rows.append("| " + " | ".join(["---"] * (max_c - min_c + 1)) + " |")

        sheet_parts.append("\n".join(rows))

        try:
            with zipfile.ZipFile(filepath, "r") as zf:
                sheet_idx = wb_formulas.sheetnames.index(sheet_name) + 1
                vml_path = f"xl/drawings/vmlDrawing{sheet_idx}.vml"
                if vml_path in zf.namelist():
                    vml_xml = zf.read(vml_path).decode("utf-8", errors="replace")
                    shapes = _extract_shapes_from_vml(vml_xml)
                    if shapes:
                        sheet_parts.append("\n### Shapes & Controls\n")
                        for shape in shapes:
                            obj_type = shape.get("object_type", "Shape")
                            name = shape.get("name", "")
                            macro = shape.get("macro", "")
                            text = shape.get("text", "")
                            desc_parts = [f"- **{obj_type}** `{name}`"]
                            if text:
                                desc_parts.append(f"text: \"{text}\"")
                            if macro:
                                desc_parts.append(f"macro: `{macro}`")
                            if desc_parts:
                                sheet_parts.append(" ".join(desc_parts))

                drawing_path = f"xl/drawings/drawing{sheet_idx}.xml"
                if drawing_path in zf.namelist():
                    drawing_xml = zf.read(drawing_path).decode("utf-8", errors="replace")
                    draw_shapes = _extract_shapes_from_drawing(drawing_xml)
                    if draw_shapes:
                        sheet_parts.append("\n### Drawing Objects\n")
                        for ds in draw_shapes:
                            name = ds.get("name", "")
                            text = ds.get("text", "")
                            macro = ds.get("macro", "")
                            desc = f"- `{name}`"
                            if text:
                                desc += f" — text: \"{text}\""
                            if macro:
                                desc += f" — macro: `{macro}`"
                            sheet_parts.append(desc)
        except (zipfile.BadZipFile, KeyError):
            pass

        parts.append("\n".join(sheet_parts))

    all_images = _extract_images(filepath)
    if all_images:
        parts.append("\n## Embedded Images\n")
        for img in all_images:
            parts.append(f"- `{img['path']}` ({img['size_bytes']:,} bytes, {img['extension']})")

    vba_section = _extract_vba(filepath)
    if vba_section:
        parts.append(vba_section)

    return "\n\n---\n\n".join(parts)
